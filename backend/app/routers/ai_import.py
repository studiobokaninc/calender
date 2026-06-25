"""
AIスマート取り込み (cmd_529 / MVP) — 抽出専用エンドポイント。

設計書正本: docs/ai_smart_import_design_2026-06-18.md

役割:
  テキストを受け取り、Claude API で「Task / Event / unknown」の種別判定 +
  フィールド抽出を行い、構造化JSON を返す。
  ★ 本EPは抽出のみ。DB へ一切保存しない（保存は FE 確認後に既存
     POST /api/tasks ・ /api/events で行う）。

安全方針:
  - ANTHROPIC_API_KEY は環境変数からのみ取得（実値はコード/コメントに残さない）。
  - 入力テキストは max_chars で上限を設けて保護。
  - anthropic SDK はハンドラ内で遅延 import（未導入環境でもアプリ起動を壊さない）。
  - 認証は通常ログインの get_current_user を利用（通常認証経路は変更しない）。
"""
import csv
import io
import os
import json
import logging
from typing import Optional, Any, Dict, List

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from pydantic import BaseModel, Field

from .. import models
from ..security import get_current_user

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/ai-import", tags=["ai_import"])

# 既定モデルは精度優先で Opus 4.8。量産時は環境変数 AI_IMPORT_MODEL で
# claude-sonnet-4-6 等へ切替可能（claude-api スキル 2026-06 準拠の正式モデルID）。
DEFAULT_MODEL = "claude-opus-4-8"
MAX_CHARS_HARD_LIMIT = 8000  # max_chars 指定がこれを超えても強制的に抑える上限

# structured outputs (output_config.format) 用の JSON Schema。
# additionalProperties:false / enum / null 型のみ使用（数値・文字長制約は非対応のため使わない）。
EXTRACTION_SCHEMA: Dict[str, Any] = {
    "type": "object",
    "additionalProperties": False,
    "required": ["kind", "confidence", "notes", "payload"],
    "properties": {
        "kind": {"type": "string", "enum": ["task", "event", "unknown"]},
        "confidence": {"type": "number"},
        "notes": {"type": "string"},
        "payload": {
            "type": "object",
            "additionalProperties": False,
            "required": [
                "title_or_name",
                "description",
                "event_type",
                "task_priority",
                "start",
                "end_or_due",
                "location",
            ],
            "properties": {
                "title_or_name": {"type": ["string", "null"]},
                "description": {"type": ["string", "null"]},
                "event_type": {
                    "type": ["string", "null"],
                    "enum": ["Meeting", "Deadline", "Milestone", "Workshop", "Generic", "Task", None],
                },
                "task_priority": {
                    "type": ["string", "null"],
                    "enum": ["HIGH", "MEDIUM", "LOW", None],
                },
                "start": {"type": ["string", "null"]},
                "end_or_due": {"type": ["string", "null"]},
                "location": {"type": ["string", "null"]},
            },
        },
    },
}

SYSTEM_PROMPT = (
    "あなたはスケジュール管理アプリの取り込みアシスタントです。"
    "ユーザーが貼り付けた自由形式テキストを読み、それが『タスク(task)』か『予定(event)』か"
    "『判別不能(unknown)』かを判定し、該当フィールドを抽出します。"
    "規則: (1)不明な項目は必ず null にし、推測で埋めない。"
    "(2)event_type は Meeting/Deadline/Milestone/Workshop/Generic/Task のいずれか、"
    "task_priority は HIGH/MEDIUM/LOW のいずれか、判断できなければ null。"
    "(3)title_or_name は最も中心的な件名/タスク名。"
    "(4)日付・時刻は分かる範囲で start / end_or_due に ISO8601 風文字列で入れる(不明なら null)。"
    "(5)confidence は判定の確信度を 0〜1 で返す。"
    "(6)notes に判断根拠や曖昧な点を簡潔に記す。"
)


class AIImportRequest(BaseModel):
    text: str = Field(..., description="取り込み対象の自由形式テキスト")
    max_chars: int = Field(default=2000, description="入力テキストの文字数上限（保護用）")


@router.post("/parse")
def parse_import_text(
    body: AIImportRequest,
    current_user: models.User = Depends(get_current_user),
) -> Dict[str, Any]:
    """
    テキストから種別判定+フィールド抽出を行い、構造化JSON を返す（保存しない）。
    """
    text = (body.text or "").strip()
    if not text:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="text は必須です。")

    # 入力上限で保護（大量テキスト対策）。max_chars はハード上限でさらに抑制。
    limit = min(max(body.max_chars, 1), MAX_CHARS_HARD_LIMIT)
    clipped = text[:limit]

    # anthropic SDK は遅延 import（未導入環境でもアプリ起動・他EPを壊さない）
    try:
        import anthropic  # type: ignore
    except Exception:
        logger.error("anthropic SDK 未導入: pip install anthropic が必要です。")
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="AI取り込み機能は未設定です（anthropic SDK が導入されていません）。",
        )

    if not os.getenv("ANTHROPIC_API_KEY"):
        logger.error("ANTHROPIC_API_KEY 未設定。")
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="AI取り込み機能は未設定です（APIキー未設定）。",
        )

    model = os.getenv("AI_IMPORT_MODEL", DEFAULT_MODEL)

    try:
        client = anthropic.Anthropic()  # APIキーは環境変数から自動解決
        response = client.messages.create(
            model=model,
            max_tokens=1024,
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": clipped}],
            output_config={
                "effort": "low",  # 分類+抽出は軽量タスク。レイテンシ優先。
                "format": {"type": "json_schema", "schema": EXTRACTION_SCHEMA},
            },
        )
    except Exception as e:  # noqa: BLE001 — 外部API障害をユーザー向けに変換
        logger.error("AI取り込み呼び出し失敗: %s", type(e).__name__)
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail="AI取り込みの実行に失敗しました。時間をおいて再試行してください。",
        )

    # output_config.format により最初の text ブロックは スキーマ準拠 JSON
    raw: Optional[str] = None
    for block in response.content:
        if getattr(block, "type", None) == "text":
            raw = block.text
            break
    if not raw:
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail="AI応答の解析に失敗しました（空応答）。",
        )

    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        logger.error("AI応答のJSONパース失敗。")
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail="AI応答の解析に失敗しました（JSON不正）。",
        )

    # 抽出結果のみ返す。保存は行わない。
    return data


# ---------------------------------------------------------------------------
# 一括抽出 (parse-bulk): 議事録・メモ等から複数アイテムを抽出
# ---------------------------------------------------------------------------

BULK_EXTRACTION_SCHEMA: Dict[str, Any] = {
    "type": "object",
    "additionalProperties": False,
    "required": ["items"],
    "properties": {
        "items": {
            "type": "array",
            "items": {
                "type": "object",
                "additionalProperties": False,
                "required": ["kind", "confidence", "notes", "payload"],
                "properties": {
                    "kind": {"type": "string", "enum": ["task", "event", "unknown"]},
                    "confidence": {"type": "number"},
                    "notes": {"type": "string"},
                    "payload": {
                        "type": "object",
                        "additionalProperties": False,
                        "required": [
                            "title_or_name",
                            "description",
                            "event_type",
                            "task_priority",
                            "start",
                            "end_or_due",
                            "location",
                        ],
                        "properties": {
                            "title_or_name": {"type": ["string", "null"]},
                            "description": {"type": ["string", "null"]},
                            "event_type": {
                                "type": ["string", "null"],
                                "enum": ["Meeting", "Deadline", "Milestone", "Workshop", "Generic", "Task", None],
                            },
                            "task_priority": {
                                "type": ["string", "null"],
                                "enum": ["HIGH", "MEDIUM", "LOW", None],
                            },
                            "start": {"type": ["string", "null"]},
                            "end_or_due": {"type": ["string", "null"]},
                            "location": {"type": ["string", "null"]},
                        },
                    },
                },
            },
        }
    },
}

BULK_SYSTEM_PROMPT = (
    "あなたはスケジュール管理アプリの取り込みアシスタントです。"
    "ユーザーが貼り付けた自由形式テキスト（議事録・メモ・箇条書き等）を読み、"
    "含まれているタスクや予定をすべて抽出し、それぞれについて種別判定と項目抽出を行います。"
    "規則: (1)不明な項目は必ず null にし、推測で埋めない。"
    "(2)event_type は Meeting/Deadline/Milestone/Workshop/Generic/Task のいずれか。"
    "(3)task_priority は HIGH/MEDIUM/LOW のいずれか、判断できなければ null。"
    "(4)confidence は各アイテムの判定確信度 0〜1。"
    "(5)notes に判断根拠や曖昧な点を簡潔に記す。"
    "(6)抽出できるアイテムがない場合は items を空配列 [] で返す。"
)


class AIImportBulkRequest(BaseModel):
    text: str = Field(..., description="取り込み対象のテキスト（議事録・メモ等）")
    max_chars: int = Field(default=4000, description="入力テキストの文字数上限（保護用）")


@router.post("/parse-bulk")
def parse_import_bulk(
    body: AIImportBulkRequest,
    current_user: models.User = Depends(get_current_user),
) -> Dict[str, Any]:
    """
    テキストから複数のタスク/イベント候補を一括抽出する（保存しない）。
    議事録・箇条書きメモ等から複数アイテムをまとめて取り込む際に使用。
    ★ 本EPは抽出のみ。DBへ一切保存しない。
    """
    text = (body.text or "").strip()
    if not text:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="text は必須です。")

    limit = min(max(body.max_chars, 1), MAX_CHARS_HARD_LIMIT * 2)  # 一括は上限2倍
    clipped = text[:limit]

    try:
        import anthropic  # type: ignore
    except Exception:
        logger.error("anthropic SDK 未導入: pip install anthropic が必要です。")
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="AI取り込み機能は未設定です（anthropic SDK が導入されていません）。",
        )

    if not os.getenv("ANTHROPIC_API_KEY"):
        logger.error("ANTHROPIC_API_KEY 未設定。")
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="AI取り込み機能は未設定です（APIキー未設定）。",
        )

    model = os.getenv("AI_IMPORT_MODEL", DEFAULT_MODEL)

    try:
        client = anthropic.Anthropic()
        response = client.messages.create(
            model=model,
            max_tokens=2048,
            system=BULK_SYSTEM_PROMPT,
            messages=[{"role": "user", "content": clipped}],
            output_config={
                "effort": "low",
                "format": {"type": "json_schema", "schema": BULK_EXTRACTION_SCHEMA},
            },
        )
    except Exception as e:
        logger.error("AI一括取り込み呼び出し失敗: %s", type(e).__name__)
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail="AI取り込みの実行に失敗しました。時間をおいて再試行してください。",
        )

    raw: Optional[str] = None
    for block in response.content:
        if getattr(block, "type", None) == "text":
            raw = block.text
            break
    if not raw:
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail="AI応答の解析に失敗しました（空応答）。",
        )

    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        logger.error("AI一括応答のJSONパース失敗。")
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail="AI応答の解析に失敗しました（JSON不正）。",
        )

    # 抽出結果のみ返す。保存は行わない。
    return data


# ---------------------------------------------------------------------------
# ファイルアップロード取り込み (parse-file): txt/md/csv/xlsx/pdf 対応
# ---------------------------------------------------------------------------

def _run_parse_llm(text: str) -> Dict[str, Any]:
    """テキストをClaude APIに投げて種別判定+抽出を行う(/parse相当)。"""
    clipped = text[:MAX_CHARS_HARD_LIMIT]

    try:
        import anthropic  # type: ignore
    except Exception:
        logger.error("anthropic SDK 未導入: pip install anthropic が必要です。")
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="AI取り込み機能は未設定です（anthropic SDK が導入されていません）。",
        )

    if not os.getenv("ANTHROPIC_API_KEY"):
        logger.error("ANTHROPIC_API_KEY 未設定。")
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="AI取り込み機能は未設定です（APIキー未設定）。",
        )

    model = os.getenv("AI_IMPORT_MODEL", DEFAULT_MODEL)

    try:
        client = anthropic.Anthropic()
        response = client.messages.create(
            model=model,
            max_tokens=1024,
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": clipped}],
            output_config={
                "effort": "low",
                "format": {"type": "json_schema", "schema": EXTRACTION_SCHEMA},
            },
        )
    except Exception as e:  # noqa: BLE001
        logger.error("AI取り込み呼び出し失敗: %s", type(e).__name__)
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail="AI取り込みの実行に失敗しました。時間をおいて再試行してください。",
        )

    raw: Optional[str] = None
    for block in response.content:
        if getattr(block, "type", None) == "text":
            raw = block.text
            break
    if not raw:
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail="AI応答の解析に失敗しました（空応答）。",
        )

    try:
        data = json.loads(raw)
    except json.JSONDecodeError:
        logger.error("AI応答のJSONパース失敗。")
        raise HTTPException(
            status_code=status.HTTP_502_BAD_GATEWAY,
            detail="AI応答の解析に失敗しました（JSON不正）。",
        )

    return data


@router.post("/parse-file")
async def parse_import_file(
    file: UploadFile = File(...),
    current_user: models.User = Depends(get_current_user),
) -> Dict[str, Any]:
    """
    ファイルアップロードを受け取り、形式に応じてテキスト抽出後にAI解析を行う（保存しない）。
    対応形式: txt / md / csv / xlsx / pdf。画像はOCR未対応（Phase2予定）。
    """
    filename = file.filename or ""
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

    if ext in ("txt", "md"):
        raw_bytes = await file.read()
        try:
            text = raw_bytes.decode("utf-8")
        except UnicodeDecodeError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="テキストファイルのUTF-8デコードに失敗しました。",
            )

    elif ext == "csv":
        raw_bytes = await file.read()
        try:
            csv_text = raw_bytes.decode("utf-8")
        except UnicodeDecodeError:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="CSVファイルのUTF-8デコードに失敗しました。",
            )
        reader = csv.reader(io.StringIO(csv_text))
        rows = [",".join(row) for row in reader]
        text = "\n".join(rows)

    elif ext == "xlsx":
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError:
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail="openpyxl未インストール。pip install openpyxl で導入してください。",
            )
        raw_bytes = await file.read()
        wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
        lines: List[str] = []
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                line = "\t".join(str(c) if c is not None else "" for c in row)
                if line.strip():
                    lines.append(line)
        text = "\n".join(lines)

    elif ext == "pdf":
        try:
            import pdfplumber  # type: ignore
        except ImportError:
            raise HTTPException(
                status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
                detail="pdfplumber未インストール。pip install pdfplumber で導入してください。段階導入推奨。",
            )
        raw_bytes = await file.read()
        with pdfplumber.open(io.BytesIO(raw_bytes)) as pdf:
            pages_text = [page.extract_text() or "" for page in pdf.pages]
        text = "\n".join(pages_text)

    elif ext in ("png", "jpg", "jpeg", "webp"):
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="画像はOCR未対応です（Phase2で対応予定）。",
        )

    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="非対応形式です。対応: txt / md / csv / xlsx / pdf",
        )

    text = text.strip()
    if not text:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="ファイルからテキストを抽出できませんでした（空または読み取り不可）。",
        )

    return _run_parse_llm(text)
